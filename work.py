from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

#-------------------Basic commands-------------------
wb = load_workbook('file name here.xlsx') #loads excel file within folder or locate file
ws = wb.active #activates file

print(ws['A1'].value) #prints value at specified point
ws['A1'].value = "Hello" #changes value at specified point or sheet

wb.save('file name here.xlsx') #saves work

#print(wb.sheetnames) #prints names of all sheets
#print(wb[Sheet1]) #prints name of specified sheet
#wb.create_sheet("Sheet2") #creates new sheet


#-------------------Example 1-------------------
wb = Workbook() #creates new workbook
wb = wb.active
ws.title = "Data" #sets title

#ws.append(['Hi', ',', 'my', 'name', 'is', 'Brandon']) #appends these to cells

for row in range(1, 11): #cycles through each row
    for col in range(0,4): #cycles through each column of each row
        #char = chr(65 + col) #get column letters manually
        char = get_column_letter #or use this function
        #print(ws[char + str(row)].value) #prints all values in sheet
        ws[char + str(row)] = char + str(row) #or change the cells to have their corresponding name in them

wb.save('Data.xlsx')


#-------------------Example 2-------------------
wb = Workbook()
wb = wb.active
ws.title = "Data"

#ws.merge_cells("A1:F3") #same formatting to merge as in excel
#ws.unmerge_cells("A1:F3") #unmerges but loses initial data

#ws.insert_rows(5) #inserts an empty row after row 5
#ws.delete_rows(5) #deletes row 

#ws.insert_cols(3) #inserts column at specified spot
#ws.delete_cols(3) #deletes column

ws.move_range("A1:F3", rows=2, cols=3) #translates range

wb.save('Data.xlsx')

#-------------------Example 3-------------------
data = { #data to be used
	"Brandon": {
		"math": 91,
		"science": 85,
		"english": 84,
		"gym": 100
	},
	"Alton": {
		"math": 55,
		"science": 72,
		"english": 87,
		"gym": 95
	},
	"Tavin": {
		"math": 100,
		"science": 45,
		"english": 75,
		"gym": 92
	},
	"Brad": {
		"math": 30,
		"science": 25,
		"english": 45,
		"gym": 100
	},
	"Kaleb": {
		"math": 100,
		"science": 100,
		"english": 100,
		"gym": 60
	}
}

wb = Workbook()
ws = wb.active
ws.title = "Grades"

headings = ['Name'] + list(data['Brandon'].keys()) #sets header
ws.append(headings) #appends header for everyone

for person in data: #loops through people
	grades = list(data[person].values()) #attributes grades to the values
	ws.append([person] + grades) #appends for everyone

for col in range(2, len(data['Brandon']) + 2):
	char = get_column_letter(col)
	ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}" #averages scores

for col in range(1, 6):
	ws[get_column_letter(col) + '1'].font = Font(bold=True, color="0099CCFF") #changes font style for headers

wb.save("NewGrades.xlsx")